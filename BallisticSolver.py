import xlrd
from openpyxl import load_workbook

def RoundedData(file):
	#openpyxl
	wb1 = load_workbook(file)
	ws1 = wb1.active
	
	#xlrd
	wb = xlrd.open_workbook(file)
	sheet = wb.sheet_by_index(0)
	
	for i in range(2, sheet.nrows):
		ws1.cell(row = (i+1), column = 8).value = round(2.3 * sheet.cell_value(i, 6))
		
		if i == (sheet.nrows):
			continue
			
	wb1.save(file)
	
file = r"C:\Users\Nathan Nangle\Desktop\Lab 4\weather.xlsx"
RoundedData(file)